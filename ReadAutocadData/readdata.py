'''
Программа считывает выделенные позиции в пространстве модели Autocad? выбирает среди них
поз "AcDbBlockReference" "Мультивыноска v1.1" (блок "Мультивыноска v1.1") и
поз. "AcDbMLeader" (Стандартная выноска Autocad) и выводит их содержимое в файл exls

'''
import copy
import os
import pathlib
import win32com.client
from Parameters import Parameters

app = win32com.client.Dispatch("AutoCAD.Application")
aDoc = app.ActiveDocument
msp = aDoc.ModelSpace
sset = aDoc.PickfirstSelectionSet

def find_and_read_parameters(sset):
    parameter_list = []
    list_poz = []
    for acad_obj in sset:
        if acad_obj.EntityName == "AcDbBlockReference":
            if acad_obj.EffectiveName == "parameters":
                parameter_list.append(acad_obj)
            elif acad_obj.EffectiveName == "Мультивыноска v1.1":
                str1 = acad_obj.GetAttributes()[0].TextString
                str2 = acad_obj.GetAttributes()[1].TextString
                print(str1, str2)
                list_poz.append(f'{str1}\\P{str2}' if str2 else str1)
        elif acad_obj.EntityName == "AcDbMLeader":
            str1 = acad_obj.TextString
            print(str1)
            list_poz.append(str1)

    if len(parameter_list) == 1:
        temp = Parameters({atr_data.TagString: atr_data.TextString for atr_data in parameter_list.pop(0).GetAttributes()})
        list_poz = {temp.level:[(temp.multiple_data, list_poz)]}
        temp.add_list_poz(list_poz)
        return temp

    else:
        raise ValueError(f'В выделеном фрагменте количество блоков параметров не равно 1. Количество найденных блоков {len(parameter_list)}')



def read_autocad_selection(EntityName=("AcDbBlockReference", "AcDbMLeader"), EffectiveName="Мультивыноска v1.1") -> list:

    sset = aDoc.PickfirstSelectionSet
    list_poz = []
    for t in sset:
        if t.EntityName in EntityName[0] and t.EffectiveName == EffectiveName:
            str1 = t.GetAttributes()[0].TextString
            str2 = t.GetAttributes()[1].TextString
            print(str1, str2)
            list_poz.append((str1, str2))
        elif t.EntityName == EntityName[1]:
            str1 = t.TextString
            str2 = ""
            print(str1, str2)
            list_poz.append((str1, str2))
    return list_poz


def read_path_file(path, filename, newPath=""):
    if newPath != "":
        try:
            with open(filename, "w", encoding='utf-8') as f:
                f.write(newPath)
            with open(filename, "r", encoding='utf-8') as f:
                path = f.readline()
        except FileNotFoundError:
            with open(filename, "w", encoding='utf-8') as f1:
                f1.write(newPath)
                path = newPath
        except PermissionError:
            print(f"Нет доступа к файлу {filename}")
            input("Для выхода нажмите Enter:")
        except:
            print(f"Ошибка файла {filename}")
            input("Для выхода нажмите Enter:")
    else:
        try:
            with open(filename, "r", encoding='utf-8') as f:
                path = f.readline()
        except FileNotFoundError:
            with open(filename, "w", encoding='utf-8') as f1:
                f1.write(str(pathlib.Path.cwd()))
                path = str(pathlib.Path.cwd())
        except PermissionError:
            print(f"Нет доступа к файлу {filename}")
            input("Для выхода нажмите Enter:")
        except:
            print(f"Ошибка файла {filename}")
            input("Для выхода нажмите Enter:")

    return path


def create_xlsx():
    import openpyxl
    workPath = read_path_file(pathlib.Path.cwd(), "path.txt")
    print(f"Текущая рабочая папка сохранения {workPath}")
    workPath = input("Введите путь к папке сохранения :")
    workPath = read_path_file(pathlib.Path.cwd(), "path.txt", workPath)
    # print(f"{workPath=}")
    filename = input("Введите имя файла :")
    if filename == "":
        filename = "output.xlsx"
    else:
        if ".xlsx" not in filename:
            filename = filename + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    active_row = 1
    for string in read_autocad_selection():
        if string[0]:
            ws.cell(active_row, 1, string[0])
            ws.cell(active_row, 2, string[1])
            active_row += 1

    try:
        os.chdir(workPath)
        wb.save(filename)
    except PermissionError:
        print("\033[31m")  # Смена цвета текста на красный
        print(
            f' ВНИМАНИЕ!!! Файл "{filename}" в папке "{workPath}" открыт. Закройте файл "{str(filename)}" и перезапустите программу.')
        print("\033[0m")  # Возврат цвета текста по умолчанию
        input()
        exit()
    print(
        f"Работа программы успешно завершена.\nОбработано {active_row - 1} {'выносок и блоков' if active_row - 1 >= 5 else 'выноски и блока'} \nСоздан файл {filename} в папке '{workPath}'")
    input("Для выхода нажмите Enter")


if __name__ == '__main__':
    const_element = find_and_read_parameters(sset)
    const_element1 = copy.deepcopy(const_element)
    print(const_element == const_element1)
    # const_element1.constr_name = 'Диафрагма Д1'
    print(const_element == const_element1)
    const_element += const_element1
    for i in dir(const_element):
        if "__" not in i:
            print(i, const_element.__getattribute__(i), type(const_element.__getattribute__(i)), sep=" --> ")


